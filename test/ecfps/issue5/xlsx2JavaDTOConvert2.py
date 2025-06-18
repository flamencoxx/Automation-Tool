import openpyxl
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple
import re
from datetime import datetime


@dataclass
class ExcelConfig:
    """Excel读取配置"""
    file_path: str
    header_row: int = 25  # 标题行号
    data_start_row: int = 26  # 数据起始行
    data_end_row: Optional[int] = None  # 数据结束行
    column_map: Dict[str, int] = None  # 列定义映射
    mandatory_col: int = 4  # 必填字段列(M/O)
    sample_col: int = 7  # 示例值列
    remark_col: int = 8  # 备注列
    position_col: int = 6  # Position列
    logical_name_col: int = 16  # Logical Field Name列
    offset_start: int = 1  # 字段偏移量起始值(从1开始，转换为0基索引)


@dataclass
class FieldConfig:
    """字段生成配置"""
    processors: List[str]  # 启用的处理器
    type_mapping: Dict[str, str]  # 类型映射
    include_meta: bool = True  # 是否包含元数据注释
    lang: str = "zh"  # 注释语言
    formatter_mapping: Dict[str, str] = None  # 类型到formatter的映射


class DTOGenerator:
    def __init__(self, excel_config: ExcelConfig, field_config: FieldConfig):
        self.excel_config = excel_config
        self.field_config = field_config
        self._init_processors()
        self.generated_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.calculated_positions = {}

    def _clean_field_name(self, field_name: str) -> str:
        """清理字段名，移除括号及其内容"""
        if not field_name:
            return field_name
        # 移除括号及其内容，如 "SETTLEMENT_DEBIT_DT (Date)" → "SETTLEMENT_DEBIT_DT"
        return re.sub(r'\s*\([^)]*\)', '', str(field_name)).strip()

    def _init_processors(self):
        """初始化处理器"""
        self.processors = {
            "camel_case": self._camel_case,
            "length_anno": self._length_anno,
            "type_anno": self._type_anno,
            "comment": self._gen_comment,
            "mandatory": self._mandatory_mark,
            "formatter": self._formatter_anno
        }

        # 默认formatter映射
        if self.field_config.formatter_mapping is None:
            self.field_config.formatter_mapping = {
                "DATE": "DDAEDateFormatter.class",
                "TIMESTAMP": "DDAEDateTimeFormatter.class",
                "NUMERIC": "BigDecimalFormatter.class",
                "DECIMAL": "BigDecimalFormatter.class"
            }

    def _camel_case(self, name: str) -> str:
        """转换为驼峰命名"""
        return re.sub(r'_([a-z])', lambda m: m.group(1).upper(), str(name).lower())

    def _length_anno(self, length: str) -> str:
        """生成长度注解"""
        return str(length) if str(length).isdigit() else ""

    def _type_anno(self, field_type: str) -> str:
        """生成类型声明"""
        type_str = self.field_config.type_mapping.get(
            re.sub(r'\(.*\)', '', field_type), 'String'
        )
        return type_str

    def _mandatory_mark(self, flag: str) -> str:
        """必填标记"""
        return flag == "M"

    def _formatter_anno(self, field_type: str) -> str:
        """生成formatter注解"""
        base_type = re.sub(r'\(.*\)', '', field_type)
        return self.field_config.formatter_mapping.get(base_type, "")

    def _gen_comment(self, field: Dict) -> str:
        """生成完整注释块"""
        comment_lines = []

        # 基础描述
        if desc := field.get("desc"):
            comment_lines.append(f"{desc}")

        # 元数据信息
        if self.field_config.include_meta:
            meta_info = [
                f"类型: {field.get('type', '')}",
                f"长度: {field.get('length', '')}",
                f"位置: {field.get('position', '')}",
                f"必填: {'是' if field.get('mandatory') == 'M' else '否'}",
                f"逻辑字段名: {field.get('logical_name', '')}" if field.get("logical_name") else None,
                f"示例: {field.get('sample', '')}" if field.get("sample") else None,
                f"备注: {field.get('remark', '')}" if field.get("remark") else None
            ]
            comment_lines.extend([m for m in meta_info if m])

        # 生成JavaDoc格式
        if comment_lines:
            return "/**\n * " + "\n * ".join(comment_lines) + "\n */"
        return ""

    def _parse_length_value(self, cell_value):
        """解析长度值（只处理数字）"""
        if cell_value is None:
            return None
        try:
            return int(float(cell_value))
        except (ValueError, TypeError):
            return None

    def _validate_positions(self, rows: List) -> None:
        """计算字段位置（基于Length累加）"""
        calculated_position = self.excel_config.offset_start  # 从1开始
        field_info = []

        print("=== 计算字段Position ===")
        for i, row in enumerate(rows, start=self.excel_config.data_start_row):
            if not any(cell.value for cell in row):
                continue

            field_name = row[0].value
            if not field_name:
                continue

            # 获取Length
            length = self._parse_length_value(row[self.excel_config.column_map["length"]].value)
            if length is None:
                print(f"警告：行 {i} 字段 {field_name} 的Length无效，跳过")
                continue

            # 记录字段信息（使用计算的Position）
            field_info.append({
                'name': field_name,
                'position': calculated_position,
                'length': length,
                'row': i
            })

            print(
                f"字段: {field_name:20} Position: {calculated_position:3} Length: {length:3} → offset: {calculated_position - 1}")

            # 更新下一个字段的Position
            calculated_position += length

        # 将计算结果存储，供后续使用
        self.calculated_positions = {info['name']: info['position'] for info in field_info}
        print(f"\n总共处理 {len(field_info)} 个字段，最终Position: {calculated_position}")

    def generate(self) -> Tuple[List[str], List[str], Dict]:
        """生成DTO字段定义"""
        wb = openpyxl.load_workbook(self.excel_config.file_path)
        sheet = wb.active

        # 读取数据行
        rows = list(sheet.iter_rows(
            min_row=self.excel_config.data_start_row,
            max_row=self.excel_config.data_end_row or sheet.max_row
        ))

        # 计算Position（不使用Excel中的值）
        self._validate_positions(rows)

        # 统计信息
        stats = {
            "total": 0,
            "mandatory": 0,
            "types": {},
            "generated_time": self.generated_time
        }

        # 处理每个字段
        fields = []
        getters_setters = []

        print("\n=== 生成字段定义 ===")
        for row in rows:
            if not row[0].value:  # 跳过空字段名
                continue

            field_name = self._clean_field_name(row[self.excel_config.column_map["name"]].value)
            if field_name not in self.calculated_positions:
                continue

            field_data = {
                "name": field_name,
                "desc": row[self.excel_config.column_map["desc"]].value,
                "type": row[self.excel_config.column_map["type"]].value,
                "length": self._parse_length_value(row[self.excel_config.column_map["length"]].value),
                "mandatory": row[self.excel_config.mandatory_col].value,
                "position": self.calculated_positions[field_name],  # 使用计算的Position
                "logical_name": row[
                    self.excel_config.logical_name_col].value if self.excel_config.logical_name_col < len(
                    row) else None,
                "sample": row[self.excel_config.sample_col].value if self.excel_config.sample_col < len(row) else None,
                "remark": row[self.excel_config.remark_col].value if self.excel_config.remark_col < len(row) else None
            }

            if field_data["length"] is None:
                print(f"警告：跳过字段 {field_name} (Length无效)")
                continue

            # 处理字段信息
            field_name_camel = self._camel_case(field_data['name'])
            field_type = self._type_anno(field_data['type'])
            field_length = self._length_anno(field_data['length'])
            is_mandatory = self._mandatory_mark(field_data['mandatory'])
            formatter = self._formatter_anno(field_data['type'])
            comment = self._gen_comment(field_data)

            # 生成字段定义
            field_def = f"{comment}\nprivate {field_type} {field_name_camel};\n"
            fields.append(field_def)

            # 生成getter注解（Position从1开始，offset从0开始，所以减1）
            offset = field_data['position'] - 1
            getter_anno = f"@Field(offset = {offset}, length = {field_length}"
            if formatter:
                getter_anno += f", formatter = {formatter}"
            getter_anno += ")"

            # 生成getter和setter
            getter = f"{comment}\n{getter_anno}\npublic {field_type} get{field_name_camel[0].upper()}{field_name_camel[1:]}() {{\n"
            getter += f"    return this.{field_name_camel};\n}}\n"

            setter = f"public void set{field_name_camel[0].upper()}{field_name_camel[1:]}({field_type} {field_name_camel}) {{\n"
            setter += f"    this.{field_name_camel} = {field_name_camel};\n}}\n"

            getters_setters.append(getter + setter)

            # 更新统计
            stats["total"] += 1
            if is_mandatory:
                stats["mandatory"] += 1
            base_type = field_data["type"].split("(")[0]
            stats["types"].setdefault(base_type, 0)
            stats["types"][base_type] += 1

        return fields, getters_setters, stats


if __name__ == "__main__":
    # 配置Excel读取规则
    excel_config = ExcelConfig(
        file_path="C:\\Users\\admin\\OneDrive\\WORK\\ecfps_doc\\mainland-crossboard\\SCB\\ISSUE-5\\FPDWPPC1_v2.40_20241028.xlsx",
        header_row=25,
        data_start_row=26,
        data_end_row=1413,
        column_map={
            "name": 0,  # A列: Field Name
            "desc": 1,  # B列: Description
            "type": 2,  # C列: Format (Output)
            "length": 5  # F列: Length
        },
        mandatory_col=4,  # E列: M/O
        sample_col=7,  # H列: Sample value
        remark_col=8,  # I列: Remarks
        position_col=6,  # G列: Position
        logical_name_col=16,
        offset_start=1  # Position从1开始
    )

    # 配置字段处理
    field_config = FieldConfig(
        processors=["comment", "mandatory", "length_anno", "formatter"],
        type_mapping={
            "CHAR": "String",
            "VARCHAR": "String",
            "NUMERIC": "BigDecimal",
            "DECIMAL": "BigDecimal",
            "DATE": "LocalDate",
            "TIMESTAMP": "LocalDateTime",
            "SMALLINT": "Integer",
            "BIGINT": "Long"
        },
        include_meta=True,
        lang="zh",
        formatter_mapping={
            "DATE": "DDAEDateFormatter.class",
            "TIMESTAMP": "DDAEDateTimeFormatter.class"
        }
    )

    # 执行生成
    generator = DTOGenerator(excel_config, field_config)
    fields, getters_setters, stats = generator.generate()

    # 生成文件头
    file_header = f"""/**
 * 自动生成的DTO类
 * 源文件: {excel_config.file_path}
 * 生成时间: {stats['generated_time']}
 * 字段统计: 共 {stats['total']} 个字段 (必填: {stats['mandatory']})
 * 类型分布: {', '.join(f'{k}:{v}' for k, v in stats['types'].items())}
 */
public class RemittanceDTO2 {{

"""

    # 输出结果
    with open("RemittanceDTO2.java", "w", encoding="utf-8") as f:
        f.write(file_header)
        f.write("\n".join(fields))
        f.write("\n")
        f.write("\n".join(getters_setters))
        f.write("\n}")

    print(f"\n生成完成，共处理 {stats['total']} 个字段")
    print(f"输出文件: RemittanceDTO2.java")
