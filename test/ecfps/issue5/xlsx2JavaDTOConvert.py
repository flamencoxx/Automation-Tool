import openpyxl
from dataclasses import dataclass
from typing import List, Dict, Optional, Tuple
import re
from datetime import datetime


@dataclass
class ExcelConfig:
    """Excel读取配置"""
    file_path: str
    header_row: int = 25  # 标题行号
    data_start_row: int = 26  # 数据起始行
    data_end_row: Optional[int] = None  # 数据结束行
    column_map: Dict[str, int] = None  # 列定义映射
    mandatory_col: int = 4  # 必填字段列(M/O)
    sample_col: int = 7  # 示例值列
    remark_col: int = 8  # 备注列
    position_col: int = 6  # Position列
    logical_name_col: int = 16  # Logical Field Name列
    offset_start: int = 0  # 字段偏移量起始值



@dataclass
class FieldConfig:
    """字段生成配置"""
    processors: List[str]  # 启用的处理器
    type_mapping: Dict[str, str]  # 类型映射
    include_meta: bool = True  # 是否包含元数据注释
    lang: str = "zh"  # 注释语言
    formatter_mapping: Dict[str, str] = None  # 类型到formatter的映射


class DTOGenerator:
    def __init__(self, excel_config: ExcelConfig, field_config: FieldConfig):
        self.excel_config = excel_config
        self.field_config = field_config
        self._init_processors()
        self.generated_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.current_offset = excel_config.offset_start

    def _init_processors(self):
        """初始化处理器"""
        self.processors = {
            "camel_case": self._camel_case,
            "length_anno": self._length_anno,
            "type_anno": self._type_anno,
            "comment": self._gen_comment,
            "mandatory": self._mandatory_mark,
            "formatter": self._formatter_anno
        }
        # 默认formatter映射
        if self.field_config.formatter_mapping is None:
            self.field_config.formatter_mapping = {
                "DATE": "DDAEDateFormatter.class",
                "TIMESTAMP": "DDAEDateTimeFormatter.class",
                "NUMERIC": "BigDecimalFormatter.class",
                "DECIMAL": "BigDecimalFormatter.class"
            }

    def _camel_case(self, name: str) -> str:
        """转换为驼峰命名"""
        return re.sub(r'_([a-z])', lambda m: m.group(1).upper(), str(name).lower())

    def _length_anno(self, length: str) -> str:
        """生成长度注解"""
        return str(length) if str(length).isdigit() else ""

    def _type_anno(self, field_type: str) -> str:
        """生成类型声明"""
        type_str = self.field_config.type_mapping.get(
            re.sub(r'\(.*\)', '', field_type), 'String'
        )
        return type_str

    def _mandatory_mark(self, flag: str) -> str:
        """必填标记"""
        return flag == "M"

    def _formatter_anno(self, field_type: str) -> str:
        """生成formatter注解"""
        base_type = re.sub(r'\(.*\)', '', field_type)
        return self.field_config.formatter_mapping.get(base_type, "")

    def _gen_comment(self, field: Dict) -> str:
        """生成完整注释块"""
        comment_lines = []

        # 基础描述
        if desc := field.get("desc"):
            comment_lines.append(f"{desc}")

        # 元数据信息
        if self.field_config.include_meta:
            meta_info = [
                f"类型: {field.get('type', '')}",
                f"长度: {field.get('length', '')}",
                f"位置: {field.get('position', '')}",
                f"必填: {'是' if field.get('mandatory') == 'M' else '否'}",
                f"逻辑字段名: {field.get('logical_name', '')}" if field.get("logical_name") else None,
                f"示例: {field.get('sample', '')}" if field.get("sample") else None,
                f"备注: {field.get('remark', '')}" if field.get("remark") else None
            ]
            comment_lines.extend([m for m in meta_info if m])

        # 生成JavaDoc格式
        if comment_lines:
            return "/**\n * " + "\n * ".join(comment_lines) + "\n */"
        return ""

    def generate(self) -> Tuple[List[str], Dict]:
        """生成DTO字段定义"""
        wb = openpyxl.load_workbook(self.excel_config.file_path)
        sheet = wb.active

        # 读取数据行
        rows = list(sheet.iter_rows(
            min_row=self.excel_config.data_start_row,
            max_row=self.excel_config.data_end_row or sheet.max_row
        ))

        # 统计信息
        stats = {
            "total": 0,
            "mandatory": 0,
            "types": {},
            "generated_time": self.generated_time
        }

        # 处理每个字段
        fields = []
        getters_setters = []
        for row in rows:
            if not row[0].value:  # 跳过空字段名
                continue

            field_data = {
                "name": row[self.excel_config.column_map["name"]].value,
                "desc": row[self.excel_config.column_map["desc"]].value,
                "type": row[self.excel_config.column_map["type"]].value,
                "length": row[self.excel_config.column_map["length"]].value,
                "mandatory": row[self.excel_config.mandatory_col].value,
                "position": row[self.excel_config.position_col].value,
                "logical_name": row[self.excel_config.logical_name_col].value,
                "sample": row[self.excel_config.sample_col].value,
                "remark": row[self.excel_config.remark_col].value
            }

            # 处理字段信息
            field_name = self._camel_case(field_data['name'])
            field_type = self._type_anno(field_data['type'])
            field_length = self._length_anno(field_data['length'])
            is_mandatory = self._mandatory_mark(field_data['mandatory'])
            formatter = self._formatter_anno(field_data['type'])
            comment = self._gen_comment(field_data)

            # 生成字段定义
            field_def = f"{comment}\nprivate {field_type} {field_name};\n"
            fields.append(field_def)

            # 生成getter注解
            getter_anno = f"@Field(offset = {self.current_offset}, length = {field_length}"
            if formatter:
                getter_anno += f", formatter = {formatter}"
            getter_anno += ")"

            # 生成getter
            getter = f"{comment}\n{getter_anno}\npublic {field_type} get{field_name[0].upper()}{field_name[1:]}() {{\n"
            getter += f"    return this.{field_name};\n}}\n"

            # 生成setter
            setter = f"public void set{field_name[0].upper()}{field_name[1:]}({field_type} {field_name}) {{\n"
            setter += f"    this.{field_name} = {field_name};\n}}\n"

            getters_setters.append(getter + setter)

            # 更新偏移量
            self.current_offset += int(field_length) if field_length.isdigit() else 0

            # 更新统计
            stats["total"] += 1
            if is_mandatory:
                stats["mandatory"] += 1
            stats["types"].setdefault(field_data["type"].split("(")[0], 0)
            stats["types"][field_data["type"].split("(")[0]] += 1

        return fields, getters_setters, stats


if __name__ == "__main__":
    # 配置Excel读取规则
    excel_config = ExcelConfig(
        file_path="C:\\Users\\admin\\OneDrive\\WORK\\ecfps_doc\\mainland-crossboard\\SCB\\ISSUE-5\\FPDWPPD_v2.40_20241028.xlsx",
        header_row=25,
        data_start_row=26,
        data_end_row=1489,
        column_map={
            "name": 0,  # A列: Field Name
            "desc": 1,  # B列: Description
            "type": 2,  # C列: Format (Output)
            "length": 5  # F列: Length
        },
        mandatory_col=4,  # E列: M/O
        sample_col=7,  # H列: Sample value
        remark_col=8,  # I列: Remarks
        offset_start=0  # 起始偏移量
    )

    # 配置字段处理
    field_config = FieldConfig(
        processors=["comment", "mandatory", "length_anno", "formatter"],
        type_mapping={
            "CHAR": "String",
            "VARCHAR": "String",
            "NUMERIC": "BigDecimal",
            "DECIMAL": "BigDecimal",
            "DATE": "LocalDate",
            "TIMESTAMP": "LocalDateTime",
            "SMALLINT": "Integer",
            "BIGINT": "Long"
        },
        include_meta=True,
        lang="zh",
        formatter_mapping={
            "DATE": "DDAEDateFormatter.class",
            "TIMESTAMP": "DDAEDateTimeFormatter.class"
        }
    )

    # 执行生成
    generator = DTOGenerator(excel_config, field_config)
    fields, getters_setters, stats = generator.generate()

    # 生成文件头
    file_header = f"""/**
 * 自动生成的DTO类
 * 源文件: {excel_config.file_path}
 * 生成时间: {stats['generated_time']}
 * 字段统计: 共 {stats['total']} 个字段 (必填: {stats['mandatory']})
 * 类型分布: {', '.join(f'{k}:{v}' for k, v in stats['types'].items())}
 */
public class RemittanceDTO {{\n"""

    # 输出结果
    with open("RemittanceDTO.java", "w", encoding="utf-8") as f:
        f.write(file_header)
        f.write("\n".join(fields))
        f.write("\n")
        f.write("\n".join(getters_setters))
        f.write("\n}")

    print(f"生成完成，共处理 {stats['total']} 个字段")
    print(f"输出文件: RemittanceDTO.java")
