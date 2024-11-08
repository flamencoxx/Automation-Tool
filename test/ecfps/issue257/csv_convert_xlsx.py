import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def append_csv_to_xlsx(csv_file, xlsx_file, sheet_name=None, start_column=1, start_row=1, insert_mode=False):
    # 读取CSV文件
    df = pd.read_csv(csv_file)

    # 加载现有的Excel文件
    wb = load_workbook(xlsx_file)

    # 如果没有指定sheet名称,就使用活动的sheet
    if sheet_name is None:
        ws = wb.active
    else:
        # 如果指定的sheet不存在,就创建一个新的
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

    # 将DataFrame转换为行,跳过标题行
    rows = list(dataframe_to_rows(df, index=False, header=False))

    if insert_mode:
        # 在插入模式下，先将现有内容向下移动
        max_row = ws.max_row
        max_col = ws.max_column
        for row in range(max_row, start_row - 1, -1):
            for col in range(1, max_col + 1):
                ws.cell(row=row + len(rows), column=col, value=ws.cell(row=row, column=col).value)

    # 添加新数据到工作表，从指定的列和行开始
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, value in enumerate(row, start_column):
            ws.cell(row=r_idx, column=c_idx, value=value)

    # 保存Excel文件
    wb.save(xlsx_file)
    print(f"数据已成功{'插入' if insert_mode else '添加'}到 {xlsx_file}")


if __name__ == '__main__':
    # 使用示例
    csv_file = 'C:\\Users\\admin\\Desktop\\ecfps_doc\\issue-257\\.csv'  # 替换为您的CSV文件名
    xlsx_file = 'C:\\Users\\admin\\Desktop\\ecfps_doc\\issue-257\\ecfps-configuration-list.xlsx'  # 替换为您的Excel文件名
    sheet_name = 'Sheet1'  # 可选,指定要添加数据的sheet名称
    start_column = 2  # 指定开始列（例如，2表示从B列开始）
    start_row = 1388  # 指定开始行（例如，134表示从第134行开始）
    insert_mode = True  # 设置为True时启用插入模式，False时保持原有的覆盖模式

    append_csv_to_xlsx(csv_file, xlsx_file, sheet_name, start_column, start_row, insert_mode)
