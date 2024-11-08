import openpyxl
import re


def process_excel(file_path, bank_codes, config):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    start_row = config['start_row']
    end_row = config['end_row'] if config['end_row'] != -1 else sheet.max_row
    columns_to_check = config['columns_to_check']

    # 处理每一行
    for row in range(start_row, end_row + 1):
        bank_found = False
        for col in columns_to_check:
            cell_value = str(sheet.cell(row=row, column=col).value or '')

            # 检查是否包含银行代码
            for bank, code in bank_codes.items():
                # 修改正则表达式以提高匹配的灵活性
                pattern = r'(?<![a-zA-Z0-9])' + re.escape(code) + r'(?![a-zA-Z0-9])'
                if re.search(pattern, cell_value, re.IGNORECASE):
                    sheet.cell(row=row, column=config['output_column'], value=bank)
                    bank_found = True
                    break

            if bank_found:
                break

    # 保存修改后的文件
    wb.save(file_path)
    print("文件处理完成")


if __name__ == "__main__":
    # 配置
    bank_codes = {
        'OCBC': 'ocbc',
        'SCB': 'scb',
        'MHBK': 'mhbk',
        'WLB': 'wlb',
        'DSB': 'dsb',
        'MOX': 'mox',
        'AMCM': 'amcm',
        'PBHK': 'pbhk'
    }

    config = {
        'start_row': 4,  # 开始检查的行
        'end_row': 1389,  # 结束检查的行，-1 表示到最后一行
        'columns_to_check': [3, 4],  # 要检查的列（C列和D列）
        'output_column': 8  # 输出结果的列（H列）
    }

    file_path = 'C:\\Users\\admin\\Desktop\\ecfps_doc\\issue-257\\ecfps-configuration-list.xlsx'
    process_excel(file_path, bank_codes, config)
